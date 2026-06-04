import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .schemas import MonthlySummaryResponse

def generate_excel_report(summary: MonthlySummaryResponse) -> io.BytesIO:
    """Generate a highly styled Excel workbook matching Requirement 3.6.2."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {summary.month}"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F2937")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="4B5563")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="1F2937")
    total_font = Font(name=font_family, size=11, bold=True, color="111827")
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
    total_fill = PatternFill(start_color="EEF2F6", end_color="EEF2F6", fill_type="solid") # Subtle gray
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thick_border_side = Side(border_style="medium", color="111827")
    double_border_side = Side(border_style="double", color="111827")
    
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_border_side, bottom=double_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 1. Title Block
    ws.cell(row=1, column=1, value=f"Payroll Management System").font = title_font
    ws.cell(row=2, column=1, value=f"Monthly Compensation Summary - {summary.month} (Lock State: {'LOCKED' if summary.is_locked else 'UNLOCKED'})").font = subtitle_font
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    
    # 2. Table Headers
    headers = [
        "Employee ID", 
        "Full Name", 
        "Aggregate Hours", 
        "Base Pay Subtotal", 
        "Aggregated Extra Work", 
        "Advance Reductions", 
        "Net Cash Payout"
    ]
    
    start_row = 4
    ws.row_dimensions[start_row].height = 28
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx in [1] else (align_left if col_idx == 2 else align_right)
        cell.border = Border(top=thick_border_side, bottom=thick_border_side, left=thin_border_side, right=thin_border_side)
        
    # 3. Data Rows
    current_row = start_row + 1
    for item in summary.items:
        ws.row_dimensions[current_row].height = 22
        
        c1 = ws.cell(row=current_row, column=1, value=item.employee_id)
        c1.alignment = align_center
        
        c2 = ws.cell(row=current_row, column=2, value=item.name)
        c2.alignment = align_left
        
        c3 = ws.cell(row=current_row, column=3, value=item.aggregate_hours)
        c3.number_format = '#,##0.00'
        c3.alignment = align_right
        
        c4 = ws.cell(row=current_row, column=4, value=item.base_pay_subtotal)
        c4.number_format = '"₹"#,##0.00'
        c4.alignment = align_right
        
        c5 = ws.cell(row=current_row, column=5, value=item.aggregated_extra_work)
        c5.number_format = '"₹"#,##0.00'
        c5.alignment = align_right
        
        c6 = ws.cell(row=current_row, column=6, value=item.advance_reductions)
        c6.number_format = '"₹"#,##0.00'
        c6.alignment = align_right
        
        c7 = ws.cell(row=current_row, column=7, value=item.net_cash_payout)
        c7.number_format = '"₹"#,##0.00'
        c7.alignment = align_right
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = border_data
            
        current_row += 1
        
    # 4. Totals Row
    ws.row_dimensions[current_row].height = 24
    ws.cell(row=current_row, column=1, value="TOTALS").alignment = align_center
    ws.cell(row=current_row, column=2, value="")
    ws.cell(row=current_row, column=3, value=summary.total_hours).number_format = '#,##0.00'
    ws.cell(row=current_row, column=4, value=summary.total_base_pay).number_format = '"₹"#,##0.00'
    ws.cell(row=current_row, column=5, value=summary.total_extra_work).number_format = '"₹"#,##0.00'
    ws.cell(row=current_row, column=6, value=summary.total_advances).number_format = '"₹"#,##0.00'
    ws.cell(row=current_row, column=7, value=summary.total_net_payout).number_format = '"₹"#,##0.00'
    
    for col_idx in range(1, 8):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border_total
        if col_idx >= 3:
            cell.alignment = align_right
            
    # Auto-fit Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Skip the title row when calculating column lengths
        for cell in col:
            if cell.row > 3 and cell.value:
                # Add formatting characters estimate to size
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Specific adjustment for Full Name
    ws.column_dimensions['B'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(summary: MonthlySummaryResponse) -> io.BytesIO:
    """Generate a clean, print-ready PDF containing registry summaries and individual breakdown vouchers."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for Premium Look
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E1B4B') # Dark Navy/Indigo
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563')
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceBefore=12,
        spaceAfter=6
    )

    voucher_h1_style = ParagraphStyle(
        'VoucherH1',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#312E81'),
        alignment=1 # Center
    )

    label_style = ParagraphStyle(
        'TableLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#374151')
    )

    value_style = ParagraphStyle(
        'TableValue',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1F2937')
    )

    value_right_style = ParagraphStyle(
        'TableValueRight',
        parent=value_style,
        alignment=2 # Right
    )

    value_bold_right_style = ParagraphStyle(
        'TableValueBoldRight',
        parent=label_style,
        alignment=2 # Right
    )

    story = []
    
    # --- PAGE 1: MONTHLY REGISTER SUMMARY ---
    story.append(Paragraph("Payroll Management System", title_style))
    story.append(Paragraph(f"Monthly Summary Ledger &mdash; {summary.month}", subtitle_style))
    story.append(Paragraph(f"Lock Status: <b>{'LOCKED (Immutable)' if summary.is_locked else 'UNLOCKED (Draft)'}</b>", subtitle_style))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Accounting Register Summary", h2_style))
    story.append(Spacer(1, 5))
    
    # Table data
    table_data = [[
        Paragraph("<b>Emp ID</b>", label_style),
        Paragraph("<b>Full Name</b>", label_style),
        Paragraph("<b>Hours</b>", label_style),
        Paragraph("<b>Base Subtotal</b>", label_style),
        Paragraph("<b>Extra Work</b>", label_style),
        Paragraph("<b>Advances</b>", label_style),
        Paragraph("<b>Net Payout</b>", label_style),
    ]]
    
    for item in summary.items:
        table_data.append([
            Paragraph(item.employee_id, value_style),
            Paragraph(item.name, value_style),
            Paragraph(f"{item.aggregate_hours:,.2f}", value_right_style),
            Paragraph(f"Rs. {item.base_pay_subtotal:,.2f}", value_right_style),
            Paragraph(f"Rs. {item.aggregated_extra_work:,.2f}", value_right_style),
            Paragraph(f"Rs. {item.advance_reductions:,.2f}", value_right_style),
            Paragraph(f"Rs. {item.net_cash_payout:,.2f}", value_bold_right_style),
        ])
        
    # Totals Row
    table_data.append([
        Paragraph("<b>TOTALS</b>", label_style),
        Paragraph("", value_style),
        Paragraph(f"<b>{summary.total_hours:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>Rs. {summary.total_base_pay:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>Rs. {summary.total_extra_work:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>Rs. {summary.total_advances:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>Rs. {summary.total_net_payout:,.2f}</b>", value_bold_right_style),
    ])
    
    # Render main summary table
    summary_table = Table(table_data, colWidths=[0.8*inch, 1.8*inch, 0.8*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F3F4F6')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 1.5, colors.HexColor('#4F46E5')),
        ('LINEBELOW', (0,1), (-1,-2), 0.5, colors.HexColor('#E5E7EB')),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#111827')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F9FAFB')),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    story.append(Paragraph("Cycle Commitments Ledger verification seal is attached to individual payout vouchers.", subtitle_style))
    story.append(PageBreak())
    
    # --- PAGES 2+: INDIVIDUAL BREAKDOWN VOUCHERS ---
    for item in summary.items:
        voucher = []
        voucher.append(Paragraph("PAYROLL RECEIPT & BREAKDOWN VOUCHER", voucher_h1_style))
        voucher.append(Spacer(1, 10))
        
        # Meta info
        meta_data = [
            [Paragraph("<b>Employee ID:</b>", label_style), Paragraph(item.employee_id, value_style),
             Paragraph("<b>Payroll Cycle:</b>", label_style), Paragraph(summary.month, value_style)],
            [Paragraph("<b>Employee Name:</b>", label_style), Paragraph(item.name, value_style),
             Paragraph("<b>Payment Status:</b>", label_style), Paragraph("PAID (Locked)" if summary.is_locked else "PENDING (Draft)", value_style)],
        ]
        meta_table = Table(meta_data, colWidths=[1.5*inch, 2.0*inch, 1.5*inch, 2.0*inch])
        meta_table.setStyle(TableStyle([
            ('LINEBELOW', (0,-1), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        voucher.append(meta_table)
        voucher.append(Spacer(1, 15))
        
        # Earnings/Deductions Details
        details_data = [
            [Paragraph("<b>Earnings Component</b>", label_style), Paragraph("<b>Calculation / Details</b>", label_style), Paragraph("<b>Subtotal</b>", value_bold_right_style)],
            
            [Paragraph("Daily Base Wage Earnings", value_style), 
             Paragraph(f"Aggregate Hours: {item.aggregate_hours:,.2f}", value_style), 
             Paragraph(f"Rs. {item.base_pay_subtotal:,.2f}", value_right_style)],
            
            [Paragraph("Extra Work & Task Extensions", value_style), 
             Paragraph("Standard/Custom Deliverable Bonuses", value_style), 
             Paragraph(f"Rs. {item.aggregated_extra_work:,.2f}", value_right_style)],
             
            [Paragraph("<b>Gross Earnings Subtotal</b>", label_style), 
             Paragraph("", value_style), 
             Paragraph(f"<b>Rs. {(item.base_pay_subtotal + item.aggregated_extra_work):,.2f}</b>", value_bold_right_style)],
             
            [Paragraph("<b>Deductions Component</b>", label_style), Paragraph("<b>Calculation / Details</b>", label_style), Paragraph("", value_bold_right_style)],
            
            [Paragraph("Cash Advances / Relieved Loans", value_style), 
             Paragraph("Accumulated advances during current cycle", value_style), 
             Paragraph(f"-Rs. {item.advance_reductions:,.2f}", value_right_style)],
             
            [Paragraph("<b>Total Reductions</b>", label_style), 
             Paragraph("", value_style), 
             Paragraph(f"<b>-Rs. {item.advance_reductions:,.2f}</b>", value_bold_right_style)],
             
            [Paragraph("<b>NET CASH DISBURSEMENT</b>", label_style), 
             Paragraph("", value_style), 
             Paragraph(f"<b>Rs. {item.net_cash_payout:,.2f}</b>", value_bold_right_style)]
        ]
        
        details_table = Table(details_data, colWidths=[2.2*inch, 3.3*inch, 1.5*inch])
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EEF2F6')),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#94A3B8')),
            ('LINEBELOW', (0,2), (-1,2), 0.5, colors.HexColor('#F1F5F9')),
            ('LINEBELOW', (0,3), (-1,3), 1, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,3), (-1,3), colors.HexColor('#F8FAFC')),
            ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#F1F5F9')),
            ('LINEBELOW', (0,4), (-1,4), 1, colors.HexColor('#94A3B8')),
            ('LINEBELOW', (0,5), (-1,5), 0.5, colors.HexColor('#F1F5F9')),
            ('LINEBELOW', (0,6), (-1,6), 1, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,6), (-1,6), colors.HexColor('#F8FAFC')),
            ('LINEABOVE', (0,7), (-1,7), 1.5, colors.HexColor('#4F46E5')),
            ('BACKGROUND', (0,7), (-1,7), colors.HexColor('#EEF2F6')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
        ]))
        voucher.append(details_table)
        voucher.append(Spacer(1, 30))
        
        # Signatures
        sig_data = [
            [Paragraph("Prepared By: __________________________", value_style), Paragraph("Employee Signature: __________________________", value_style)],
            [Paragraph("Super Administrator Authority", subtitle_style), Paragraph("Acknowledge of receipt of net payout", subtitle_style)]
        ]
        sig_table = Table(sig_data, colWidths=[3.5*inch, 3.5*inch])
        sig_table.setStyle(TableStyle([
            ('TOPPADDING', (0,0), (-1,0), 20),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        voucher.append(sig_table)
        
        story.append(KeepTogether(voucher))
        story.append(PageBreak())
        
    doc.build(story)
    output.seek(0)
    return output


def generate_employee_excel_report(report) -> io.BytesIO:
    """Generate a styled Excel workbook for an individual employee's monthly daily log."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report.employee_id} {report.month}"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F2937")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="4B5563")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="1F2937")
    total_font = Font(name=font_family, size=11, bold=True, color="111827")
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    total_fill = PatternFill(start_color="EEF2F6", end_color="EEF2F6", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thick_border_side = Side(border_style="medium", color="111827")
    double_border_side = Side(border_style="double", color="111827")
    
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_border_side, bottom=double_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Title Block
    ws.cell(row=1, column=1, value=f"Individual Employee Monthly Payroll Log").font = title_font
    ws.cell(row=2, column=1, value=f"Employee: {report.name} ({report.employee_id})  |  Month: {report.month}  |  Base Rate: ₹{report.hourly_rate:.2f}/hr").font = subtitle_font
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    
    # Headers
    headers = [
        "Date",
        "Status",
        "Hours Logged",
        "Hourly Rate",
        "Base Earnings",
        "Extra Work Tags & Amounts",
        "Cash Advances & Deductions",
        "Net Daily Payout"
    ]
    
    start_row = 4
    ws.row_dimensions[start_row].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx in [1, 2] else (align_left if col_idx in [6, 7] else align_right)
        cell.border = Border(top=thick_border_side, bottom=thick_border_side, left=thin_border_side, right=thin_border_side)
        
    # Data Rows
    current_row = start_row + 1
    for day in report.days:
        ws.row_dimensions[current_row].height = 22
        
        ws.cell(row=current_row, column=1, value=day.date.strftime("%Y-%m-%d")).alignment = align_center
        ws.cell(row=current_row, column=2, value=day.status).alignment = align_center
        
        c3 = ws.cell(row=current_row, column=3, value=day.hours_logged if day.status == "Present" else "")
        if day.status == "Present":
            c3.number_format = '#,##0.00'
        c3.alignment = align_right
        
        c4 = ws.cell(row=current_row, column=4, value=day.base_rate)
        c4.number_format = '"₹"#,##0.00'
        c4.alignment = align_right
        
        c5 = ws.cell(row=current_row, column=5, value=day.base_pay)
        c5.number_format = '"₹"#,##0.00'
        c5.alignment = align_right
        
        # Extra Work tags description
        extra_desc = ", ".join([f"{item.tag} (+₹{item.amount:.2f})" for item in day.extra_work_items])
        ws.cell(row=current_row, column=6, value=extra_desc or "-").alignment = align_left
        
        # Advances description
        adv_desc = ", ".join([f"Loan (-₹{item.amount:.2f})" for item in day.advance_items])
        ws.cell(row=current_row, column=7, value=adv_desc or "-").alignment = align_left
        
        c8 = ws.cell(row=current_row, column=8, value=day.net_pay)
        c8.number_format = '"₹"#,##0.00'
        c8.alignment = align_right
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = border_data
            
        current_row += 1
        
    # Totals Row
    ws.row_dimensions[current_row].height = 24
    ws.cell(row=current_row, column=1, value="TOTALS").alignment = align_center
    ws.cell(row=current_row, column=2, value="")
    
    ws.cell(row=current_row, column=3, value=report.total_hours).number_format = '#,##0.00'
    ws.cell(row=current_row, column=4, value="")
    ws.cell(row=current_row, column=5, value=report.total_base_pay).number_format = '"₹"#,##0.00'
    ws.cell(row=current_row, column=6, value=f"Bonus: ₹{report.total_extra_work:.2f}")
    ws.cell(row=current_row, column=7, value=f"Advance: ₹{report.total_advances:.2f}")
    ws.cell(row=current_row, column=8, value=report.total_net_payout).number_format = '"₹"#,##0.00'
    
    for col_idx in range(1, 9):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border_total
        if col_idx in [3, 5, 8]:
            cell.alignment = align_right
        elif col_idx in [6, 7]:
            cell.alignment = align_left
            
    # Auto Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_employee_pdf_report(report) -> io.BytesIO:
    """Generate a clean, print-ready PDF containing daily attendance log vouchers for an employee."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'EmpDocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1E1B4B')
    )
    
    subtitle_style = ParagraphStyle(
        'EmpDocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563')
    )
    
    label_style = ParagraphStyle(
        'EmpTableLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#374151')
    )
    
    value_style = ParagraphStyle(
        'EmpTableValue',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1F2937')
    )
    
    value_right_style = ParagraphStyle(
        'EmpTableValueRight',
        parent=value_style,
        alignment=2 # Right
    )
    
    value_bold_right_style = ParagraphStyle(
        'EmpTableValueBoldRight',
        parent=label_style,
        alignment=2 # Right
    )
    
    story = []
    
    # Document Header
    story.append(Paragraph("Employee Monthly Payroll Log Report", title_style))
    story.append(Spacer(1, 8))
    
    # Metadata Block Table
    meta_data = [
        [Paragraph("<b>Employee Name:</b>", label_style), Paragraph(report.name, value_style),
         Paragraph("<b>Employee ID:</b>", label_style), Paragraph(report.employee_id, value_style)],
        [Paragraph("<b>Payroll Month:</b>", label_style), Paragraph(report.month, value_style),
         Paragraph("<b>Hourly Base Rate:</b>", label_style), Paragraph(f"Rs. {report.hourly_rate:.2f}/hr", value_style)]
    ]
    meta_table = Table(meta_data, colWidths=[1.5*inch, 2.0*inch, 1.5*inch, 2.0*inch])
    meta_table.setStyle(TableStyle([
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # Table data
    table_data = [[
        Paragraph("<b>Date</b>", label_style),
        Paragraph("<b>Status</b>", label_style),
        Paragraph("<b>Hours Logged</b>", label_style),
        Paragraph("<b>Hourly Rate</b>", label_style),
        Paragraph("<b>Base Pay</b>", label_style),
        Paragraph("<b>Extra Work Bonuses</b>", label_style),
        Paragraph("<b>Advances</b>", label_style),
        Paragraph("<b>Net Payout</b>", label_style),
    ]]
    
    for day in report.days:
        extra_desc = ", ".join([f"{item.tag} (+Rs. {item.amount:.2f})" for item in day.extra_work_items]) or "-"
        adv_desc = ", ".join([f"Loan (-Rs. {item.amount:.2f})" for item in day.advance_items]) or "-"
        
        table_data.append([
            Paragraph(day.date.strftime("%b %d, %a"), value_style),
            Paragraph(day.status, value_style),
            Paragraph(f"{day.hours_logged:,.2f}" if day.status == "Present" else "-", value_right_style),
            Paragraph(f"Rs. {day.base_rate:,.2f}", value_right_style),
            Paragraph(f"Rs. {day.base_pay:,.2f}", value_right_style),
            Paragraph(extra_desc, value_style),
            Paragraph(adv_desc, value_style),
            Paragraph(f"Rs. {day.net_pay:,.2f}", value_bold_right_style),
        ])
        
    # Totals Row
    table_data.append([
        Paragraph("<b>TOTALS</b>", label_style),
        Paragraph("", value_style),
        Paragraph(f"<b>{report.total_hours:,.2f}</b>", value_bold_right_style),
        Paragraph("", value_style),
        Paragraph(f"<b>Rs. {report.total_base_pay:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>+Rs. {report.total_extra_work:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>-Rs. {report.total_advances:,.2f}</b>", value_bold_right_style),
        Paragraph(f"<b>Rs. {report.total_net_payout:,.2f}</b>", value_bold_right_style),
    ])
    
    # Render table
    ledger_table = Table(table_data, colWidths=[0.9*inch, 0.6*inch, 0.7*inch, 0.6*inch, 0.7*inch, 1.5*inch, 1.4*inch, 1.1*inch])
    ledger_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F3F4F6')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,0), 1.5, colors.HexColor('#4F46E5')),
        ('LINEBELOW', (0,1), (-1,-2), 0.5, colors.HexColor('#E5E7EB')),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#111827')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F9FAFB')),
    ]))
    
    story.append(ledger_table)
    story.append(Spacer(1, 30))
    
    # Signatures
    sig_data = [
        [Paragraph("Prepared By: __________________________", value_style), Paragraph("Employee Signature: __________________________", value_style)],
        [Paragraph("Super Administrator Authority", subtitle_style), Paragraph("Acknowledge of receipt of net payout", subtitle_style)]
    ]
    sig_table = Table(sig_data, colWidths=[3.75*inch, 3.75*inch])
    sig_table.setStyle(TableStyle([
        ('TOPPADDING', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    output.seek(0)
    return output
